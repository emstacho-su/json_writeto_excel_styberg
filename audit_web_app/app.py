from flask import Flask, render_template, request, redirect, flash
import pandas as pd
import json
from openpyxl import load_workbook, Workbook
import os

# For file save dialog
import tkinter as tk
from tkinter import filedialog

app = Flask(__name__)
app.secret_key = 'secure-key'

EXCEL_FILENAME = "audit_data.xlsx"
EXCEL_PATH = None  # Will be set dynamically on first save

SHEET_MAP = {
    "Null Hypothesis": ["work_instruction", "clause", "statistical_test", "p_value", "effect_size", "compliance"],
    "Material Evidence": ["work_instruction", "clause", "evidence_summary", "evidence_grade", "coverage"],
    "Gap Severity": ["work_instruction", "clause", "gap_severity", "gap_description"],
    "Longitudinal Tracking": ["work_instruction", "metric", "value", "longitudinal_notes"]
}



def get_excel_path():
    global EXCEL_PATH

    if EXCEL_PATH and os.path.exists(EXCEL_PATH):
        return EXCEL_PATH

    # Ask user where to save the Excel file
    root = tk.Tk()
    root.withdraw()
    file_path = filedialog.asksaveasfilename(
        defaultextension=".xlsx",
        filetypes=[("Excel files", "*.xlsx")],
        title="Select location to save audit_data.xlsx",
        initialfile=EXCEL_FILENAME
    )

    if file_path:
        # Create blank workbook with all expected sheets
        wb = Workbook()
        ws = wb.active
        ws.title = list(SHEET_MAP.keys())[0]

        # Add the remaining sheets
        for sheet_name in list(SHEET_MAP.keys())[1:]:
            wb.create_sheet(title=sheet_name)

        wb.save(file_path)
        EXCEL_PATH = file_path
        return file_path
    else:
        return None


@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        json_data = request.form.get('json_input')
        try:
            data = json.loads(json_data)

            path = get_excel_path()
            if not path:
                flash("Save cancelled. No Excel file was created.", 'error')
                return redirect('/')

            book = load_workbook(path)
            writer = pd.ExcelWriter(path, engine='openpyxl', mode='a', if_sheet_exists='overlay')
            writer.book = book

            for sheet, columns in SHEET_MAP.items():
                if sheet not in data:
                    continue

                df_new = pd.DataFrame(data[sheet])
                df_new = df_new[columns]

                if sheet in book.sheetnames:
                    df_existing = pd.read_excel(path, sheet_name=sheet)
                    df_combined = pd.concat([df_existing, df_new], ignore_index=True)
                    df_combined.to_excel(writer, sheet_name=sheet, index=False)
                else:
                    df_new.to_excel(writer, sheet_name=sheet, index=False)

            writer.save()
            writer.close()
            flash("✅ Data successfully saved to Excel file.", 'success')
        except json.JSONDecodeError:
            flash("Invalid JSON input. Please check your formatting.", 'error')
        except Exception as e:
            flash(f"An error occurred: {str(e)}", 'error')

        return redirect('/')

    return render_template('index.html')
